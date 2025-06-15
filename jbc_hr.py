# CV Analyzer for Google Drive - JBC HR AI Assistant
# Mount Google Drive, extract CV information from PDFs, and create an Excel output file

import os
import pandas as pd
import io
import PyPDF2
import openai
import re
import tempfile
import json
from datetime import datetime
import dateutil.parser
from dateutil.relativedelta import relativedelta
from google.colab import drive
from tqdm.notebook import tqdm
from googleapiclient.discovery import build
from google.colab import auth
from google.auth import default
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from google.colab import files

# Mount Google Drive
drive.mount('/content/drive')

# Set your OpenAI API key
openai_api_key = "YOUR_OPENAI_API_KEY_HERE"  # Replace with your actual API key
openai.api_key = openai_api_key

# Define the root path
root_path = "/content/drive/MyDrive/JBC_HR_AI_ASSISTANT"

# Google Drive file link base URL
DRIVE_LINK_BASE = "https://drive.google.com/file/d/"

# Function to get Google Drive file ID
def get_file_id(file_path):
    """Get Google Drive file ID from local path"""
    try:
        # Authenticate and create the Drive API client
        auth.authenticate_user()
        creds, _ = default()
        drive_service = build('drive', 'v3', credentials=creds)

        # Extract the relative path from the full path
        relative_path = os.path.relpath(file_path, "/content/drive/MyDrive")

        # Search for the file by name
        filename = os.path.basename(file_path)
        query = f"name = '{filename}' and trashed = false"

        # Execute the query
        results = drive_service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name, parents)'
        ).execute()

        items = results.get('files', [])

        if not items:
            return None

        # If multiple files have the same name, try to match the path
        if len(items) > 1:
            for item in items:
                # Get the file's complete path
                file_path_in_drive = get_file_path_in_drive(drive_service, item['id'])
                if relative_path in file_path_in_drive:
                    return item['id']

            # If no path match, return the first one
            return items[0]['id']
        else:
            return items[0]['id']
    except Exception as e:
        print(f"Error getting file ID: {str(e)}")
        return None

# Function to get file path in Drive
def get_file_path_in_drive(service, file_id):
    """Get the file path in Google Drive"""
    try:
        # Get the file metadata
        file = service.files().get(fileId=file_id, fields='name, parents').execute()

        path = [file['name']]

        # Get all parent folders
        if 'parents' in file:
            parent_id = file['parents'][0]
            while parent_id:
                parent = service.files().get(fileId=parent_id, fields='name, parents').execute()
                path.insert(0, parent['name'])

                if 'parents' in parent:
                    parent_id = parent['parents'][0]
                else:
                    parent_id = None

        return '/'.join(path)
    except Exception as e:
        print(f"Error getting file path: {str(e)}")
        return ""

# Function to create shareable link
def create_shareable_link(file_id):
    """Create a shareable link for the Google Drive file"""
    if file_id:
        return f"{DRIVE_LINK_BASE}{file_id}/view?usp=sharing"
    return "Link not available"

def extract_text_from_pdf(pdf_file):
    """Extract text content from a PDF file."""
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page_num in range(len(pdf_reader.pages)):
        text += pdf_reader.pages[page_num].extract_text()
    return text

def calculate_experience_duration(start_date_str):
    """Calculate duration between start date and current date in 'X year Y month' format."""
    try:
        # Parse the start date string
        if start_date_str == "Not found" or not start_date_str:
            return "Not found"

        # Try to parse the date with dateutil parser
        try:
            start_date = dateutil.parser.parse(start_date_str, fuzzy=True)
        except:
            # If parsing fails, try to extract month and year manually
            match = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* (\d{4})',
                             start_date_str, re.IGNORECASE)
            if match:
                month_str = match.group(1)
                year_str = match.group(2)
                # Map abbreviated month to number
                month_map = {
                    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
                    'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
                    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
                }
                month = month_map.get(month_str.lower()[:3], 1)
                year = int(year_str)
                start_date = datetime(year, month, 1)
            else:
                return "Date format not recognized"

        # Calculate the difference between the start date and current date
        current_date = datetime.now()
        delta = relativedelta(current_date, start_date)

        # Format the result as "X year Y month"
        years = delta.years
        months = delta.months

        if years == 0:
            if months == 1:
                return f"{months} month"
            else:
                return f"{months} months"
        elif years == 1:
            if months == 0:
                return "1 year"
            elif months == 1:
                return "1 year 1 month"
            else:
                return f"1 year {months} months"
        else:
            if months == 0:
                return f"{years} years"
            elif months == 1:
                return f"{years} years 1 month"
            else:
                return f"{years} years {months} months"
    except Exception as e:
        return f"Error calculating duration: {str(e)}"

def extract_field(text, field_name):
    """Extract a specific field from text response when JSON parsing fails"""
    pattern = rf"{field_name}[:\s]+(.*?)(?:\n|$|,)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return "Not found"

def extract_cv_info(cv_text):
    """Use OpenAI API to extract structured information from CV text."""

    # Get current date for calculating work experience
    current_date = datetime.now()
    current_date_str = current_date.strftime("%Y %B")

    prompt = f"""
    Extract the following information from the CV text below.
    If you cannot find a particular piece of information, respond with "Not found" for that field.

    Information to extract:
    1. Name
    2. Last Education and university
    3. Number of total year experiences
    4. Present field of experience
    5. Overall expertise area
    6. Present organization designation
    7. Research experience (any research positions, publications, or projects)
    8. Achievements (awards, recognitions, significant accomplishments)
    9. Mobile number
    10. Email address
    11. Present organization name
    12. Working experience in present organization (start date in format 'Month YYYY', e.g. 'December 2022')

    Today is {current_date_str}.

    CV Text:
    {cv_text}

    Your response MUST be a valid JSON object with ONLY the following keys:
    {{
      "name": "extracted name",
      "last_education": "extracted education and university",
      "total_experience": "total number of experiences in all organizations",
      "present_field": "present field of experience",
      "overall_expertise_area": "areas of expertise or specialization",
      "present_organization_designation": "current job title or designation",
      "research_experience": "details of research experience if any",
      "achievements": "major achievements and awards if any",
      "mobile": "extracted mobile number",
      "email": "extracted email address",
      "present_organization_name": "name of current organization",
      "working_experience_in_present_organization": "start date in format 'Month YYYY'"
    }}

    Do not include any explanation, just return the JSON object.
    """

    try:
        # Use GPT-4o model for better extraction
        model = "gpt-4o"

        response = openai.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "You are a helpful assistant that extracts structured information from CVs. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3  # Lower temperature for more consistent results
        )

        # Extract and parse the JSON response
        result = response.choices[0].message.content

        # Try to find JSON in the response
        json_match = re.search(r'(\{[\s\S]*\})', result, re.DOTALL)
        if json_match:
            json_str = json_match.group(1)
            try:
                parsed_result = json.loads(json_str)

                # Calculate work experience duration
                start_date = parsed_result.get("working_experience_in_present_organization", "Not found")
                parsed_result["working_experience_in_year_in_present_organization"] = calculate_experience_duration(start_date)

                return parsed_result
            except json.JSONDecodeError:
                print(f"Failed to parse JSON from response. Attempting alternate extraction.")

        # If extraction failed, try to create a structured response manually
        try:
            # Create a standard response manually
            start_date = extract_field(result, "working_experience_in_present_organization")
            experience_duration = calculate_experience_duration(start_date)

            # Return in the specified order
            return {
                "name": extract_field(result, "name"),
                "last_education": extract_field(result, "last_education"),
                "overall_expertise_area": extract_field(result, "overall_expertise_area"),
                "present_organization_name": extract_field(result, "present_organization_name"),
                "present_organization_designation": extract_field(result, "present_organization_designation"),
                "working_experience_in_present_organization": start_date,
                "working_experience_in_year_in_present_organization": experience_duration,
                "total_experience": extract_field(result, "total_experience"),
                "present_field": extract_field(result, "present_field"),
                "research_experience": extract_field(result, "research_experience"),
                "achievements": extract_field(result, "achievements"),
                "mobile": extract_field(result, "mobile"),
                "email": extract_field(result, "email")
            }
        except Exception as e:
            print(f"Error creating structured response: {str(e)}")
            # Last resort, try direct JSON parsing
            parsed_result = json.loads(result)

            # Calculate work experience duration
            start_date = parsed_result.get("working_experience_in_present_organization", "Not found")
            parsed_result["working_experience_in_year_in_present_organization"] = calculate_experience_duration(start_date)

            return parsed_result

    except Exception as e:
        print(f"Error extracting information: {str(e)}")
        return {
            "name": "Error",
            "last_education": "Error",
            "total_experience": "Error",
            "present_field": "Error",
            "overall_expertise_area": "Error",
            "present_organization_designation": "Error",
            "research_experience": "Error",
            "achievements": "Error",
            "present_organization_name": "Error",
            "working_experience_in_present_organization": "Error",
            "working_experience_in_year_in_present_organization": "Error",
            "mobile": "Error",
            "email": "Error"
        }

# Function to find all PDF files in the directory structure
def find_pdf_files(root_path):
    pdf_files = []
    for root, _, files in os.walk(root_path):
        for file in files:
            if file.lower().endswith('.pdf'):
                # Get relative path components
                rel_path = os.path.relpath(root, root_path)
                if rel_path == '.':
                    subfolder = ""
                else:
                    subfolder = rel_path

                full_path = os.path.join(root, file)

                # Get file ID and create shareable link
                file_id = get_file_id(full_path)
                cv_link = create_shareable_link(file_id)

                pdf_files.append({
                    'full_path': full_path,
                    'filename': file,
                    'root_folder': root_path,
                    'subfolder': subfolder,
                    'cv_link': cv_link
                })
    return pdf_files

# Main execution
def main():
    print(f"Starting CV analysis from root path: {root_path}")

    # Find all PDF files
    pdf_files = find_pdf_files(root_path)
    print(f"Found {len(pdf_files)} PDF files to process")

    # Create a list to store results
    all_results = []

    # Process each PDF file
    for pdf_info in tqdm(pdf_files, desc="Processing CVs"):
        try:
            # Open the PDF file and extract text
            with open(pdf_info['full_path'], 'rb') as file:
                # Extract text from PDF
                cv_text = extract_text_from_pdf(file)

                # Extract structured information
                cv_info = extract_cv_info(cv_text)

                # Add file information
                cv_info["cv_links"] = pdf_info['filename']  # Will be displayed and renamed later
                cv_info["root_folder"] = pdf_info['root_folder']
                cv_info["subfolder"] = pdf_info['subfolder']
                cv_info["_cv_link"] = pdf_info['cv_link']  # Temporary field for processing

                # Add to results list
                all_results.append(cv_info)

                # Print progress
                print(f"Processed: {pdf_info['filename']}")
        except Exception as e:
            print(f"Error processing {pdf_info['filename']}: {str(e)}")
            # Add error entry
            error_info = {
                "name": "Error",
                "last_education": "Error",
                "total_experience": "Error",
                "present_field": "Error",
                "overall_expertise_area": "Error",
                "present_organization_designation": "Error",
                "research_experience": "Error",
                "achievements": "Error",
                "present_organization_name": "Error",
                "working_experience_in_present_organization": "Error",
                "working_experience_in_year_in_present_organization": "Error",
                "mobile": "Error",
                "email": "Error",
                "cv_links": pdf_info['filename'],
                "root_folder": pdf_info['root_folder'],
                "subfolder": pdf_info['subfolder'],
                "_cv_link": pdf_info['cv_link']
            }
            all_results.append(error_info)

    # Create a DataFrame from all results
    df = pd.DataFrame(all_results)

    # Define the desired column order
    column_order = [
        "name",
        "last_education",
        "overall_expertise_area",
        "present_organization_name",
        "present_organization_designation",
        "working_experience_in_present_organization",
        "working_experience_in_year_in_present_organization",
        "total_experience",
        "present_field",
        "research_experience",
        "achievements",
        "mobile",
        "email",
        "cv_links",  # Renamed from filename and will be the hyperlink
        "root_folder",
        "subfolder"
    ]

    # Reorder columns (only include columns that exist)
    existing_columns = [col for col in column_order if col in df.columns]
    extra_columns = [col for col in df.columns if col not in column_order]
    df = df[existing_columns + extra_columns]

    # Ensure all data is treated as strings to avoid conversion issues
    for column in df.columns:
        df[column] = df[column].astype(str)

    # Save the DataFrame to Excel
    output_path = os.path.join(root_path, "cv_analysis_results.xlsx")

    # Remove the temporary _cv_link column before saving
    if "_cv_link" in df.columns:
        df_save = df.drop("_cv_link", axis=1)
    else:
        df_save = df.copy()

    # First save using pandas to get the basic structure
    df_save.to_excel(output_path, index=False)

    # Now, modify the Excel file to create proper hyperlinks
    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    # Find the column indices for cv_links
    header_row = worksheet[1]
    cv_links_col_idx = None

    for idx, cell in enumerate(header_row, 1):
        if cell.value == "cv_links":
            cv_links_col_idx = idx

    # If the column exists, create hyperlinks
    if cv_links_col_idx:
        for row_idx in range(2, worksheet.max_row + 1):
            # Find the corresponding link from the original DataFrame
            if row_idx - 2 < len(df) and "_cv_link" in df.columns:
                file_link = df.iloc[row_idx-2]["_cv_link"]

                # Get the filename
                cv_links_cell = worksheet.cell(row=row_idx, column=cv_links_col_idx)
                filename = cv_links_cell.value

                if filename and file_link and file_link != "Link not available":
                    # Set the hyperlink
                    cv_links_cell.hyperlink = file_link
                    cv_links_cell.value = filename

                    # Format the cell as a hyperlink (blue and underlined)
                    cv_links_cell.font = Font(color="0000FF", underline="single")

    # Save the modified workbook to a final path for download
    final_output_path = os.path.join(root_path, "cv_analysis_results_with_links.xlsx")
    workbook.save(final_output_path)

    print(f"Analysis complete! Results saved to: {final_output_path} with clickable hyperlinks")

    # Automatically download the file
    files.download(final_output_path)

    return df_save

# Run the main function
if __name__ == "__main__":
    result_df = main()
    display(result_df)  # Display the results in the notebook